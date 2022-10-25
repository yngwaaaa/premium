#!/usr/bin/env python
# coding: utf-8

# In[3]:


import AXA_func as AXA
from tqdm import tqdm
from time import sleep

import pandas as pd
import openpyxl as xl

import os
import sys
import random


# In[4]:


FILE_NAME=input('ファイル名を入力（拡張子.xlsmは除く）')
SHEET_NAME='AXA打鍵'
BASE_BAT_SIZE=5 #バッチサイズ。一度の打鍵件数

#最初にバックアップファイルがない場合に作成
if os.path.isfile(FILE_NAME+'_backup.xlsm'):
    pass
else:
    wb=xl.load_workbook(FILE_NAME+'.xlsm',keep_vba=True)
    wb.save(FILE_NAME+'_backup.xlsm')

#打鍵する行のリストの箱を用意
calc_row=list(range(1))

#打鍵、バッチ単位のループ
while len(calc_row) > 0:
    try:
        df=pd.read_excel(FILE_NAME+'.xlsm',sheet_name=SHEET_NAME)
        wb=xl.load_workbook(FILE_NAME+'.xlsm',keep_vba=True)    #ループ中に出力するために、WBも読み込んでおく
        wb.save(FILE_NAME+'_backup.xlsm')#読み込み成功したらバックアップを上書き
    except:#エラーが起きたらバックアップファイルから読み込む
        print('File load error')
        df=pd.read_excel(FILE_NAME+'_backup.xlsm',sheet_name=SHEET_NAME)
        wb=xl.load_workbook(FILE_NAME+'_backup.xlsm',keep_vba=True)    #ループ中に出力するために、WBも読み込んでおく
        wb.save(FILE_NAME+'.xlsm')

    #列を追加
    df['車有P_賠償']=''
    df['車無P_賠償']=''
    df['車有P_傷害']=''
    df['車無P_傷害']=''
    df['車有P_車両']=''
    df['車無P_車両']=''
    df['車有P_その他']=''
    df['車無P_その他']=''
    df['車両AMTエラー']=''
    df['新車保険金額エラー']=''

    ws=wb[SHEET_NAME]
    ws.cell(row=1, column=84).value = '車有P_賠償' 
    ws.cell(row=1, column=85).value = '車無P_賠償'
    ws.cell(row=1, column=86).value = '車有P_傷害' 
    ws.cell(row=1, column=87).value = '車無P_傷害'
    ws.cell(row=1, column=88).value = '車有P_車両' 
    ws.cell(row=1, column=89).value = '車無P_車両'
    ws.cell(row=1, column=90).value = '車有P_その他' 
    ws.cell(row=1, column=91).value = '車無P_その他'
    ws.cell(row=1, column=92).value = '車両AMTエラー'
    ws.cell(row=1, column=93).value = '新車保険金額エラー'

    #打鍵する行を特定
    calc_row=list()
    #数値のはいっていない行のうち、バッチサイズの行だけExcel上の行番号を取得（dfのindex+2）する。
    BAT_SIZE=BASE_BAT_SIZE + random.randint(0, 5) #並列処理時にデータファイルアクセスのタイミングをずらすために乱数を加算
    calc_row=list(df[(df['車有P'] == 'E') | (df['車有P'].isna())].index[0:BAT_SIZE]+2)
    print(calc_row)

    #打鍵を始める行に打鍵中と入力
    for j in calc_row:
        ws.cell(row=j, column=80).value = '打鍵中' 
    wb.save(FILE_NAME+'.xlsm')#いったん保存

    #####打鍵、行単位のループ#####################################  
    for i in tqdm(calc_row):
        data=df.loc[i-2,:].to_dict()
        data = AXA.AXA_func(data)#打鍵

        #結果をdfに書き込む
        df_temp=pd.DataFrame.from_dict(data, orient='index').T
        df.loc[i-2,:] = df_temp.iloc[0,:]
    #####行単位のループ終了######################################

    ########並列で実行するため、あらためて現時点の最新版のファイルを読み出して結果を追加
    try:
        wb=xl.load_workbook(FILE_NAME+'.xlsm',keep_vba=True)
        wb.save(FILE_NAME+'_backup.xlsm')#読み込み成功したらバックアップを上書き
    except:#エラーが起きたらバックアップファイルを復旧
        print('File load error')
        wb=xl.load_workbook(FILE_NAME+'_backup.xlsm',keep_vba=True)
        wb.save(FILE_NAME+'.xlsm')
    
    ws=wb[SHEET_NAME]
    for i in calc_row:
        ws.cell(row=i, column=80).value = df.loc[i-2,'車有P'] 
        ws.cell(row=i, column=81).value = df.loc[i-2,'車無P']
        ws.cell(row=i, column=82).value = df.loc[i-2,'イ割なし車有'] 
        ws.cell(row=i, column=83).value = df.loc[i-2,'イ割なし車無']
        ws.cell(row=i, column=84).value = df.loc[i-2,'車有P_賠償'] 
        ws.cell(row=i, column=85).value = df.loc[i-2,'車無P_賠償']
        ws.cell(row=i, column=86).value = df.loc[i-2,'車有P_傷害'] 
        ws.cell(row=i, column=87).value = df.loc[i-2,'車無P_傷害']
        ws.cell(row=i, column=88).value = df.loc[i-2,'車有P_車両'] 
        ws.cell(row=i, column=89).value = df.loc[i-2,'車無P_車両']
        ws.cell(row=i, column=90).value = df.loc[i-2,'車有P_その他'] 
        ws.cell(row=i, column=91).value = df.loc[i-2,'車無P_その他']
        ws.cell(row=i, column=92).value = df.loc[i-2,'車両AMTエラー']
        ws.cell(row=i, column=93).value = df.loc[i-2,'新車保険金額エラー']

    wb.save(FILE_NAME+'.xlsm')
#バッチ単位のループ終了########################################

#バックアップを削除
os.remove(FILE_NAME+'_backup.xlsm')

    

